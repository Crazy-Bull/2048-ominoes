import pickle
from state import int2state, state_sum
import os
from calculate import Expectations, rnd_spawn_int
import openpyxl
from openpyxl.drawing.image import Image

def unprocessed():
    unprocessed_list = []
    processed_list = []
    for i in range(369):
        path = f"polynominoes/{i}/sum_4.pkl"
        if not os.path.exists(path):
            unprocessed_list.append(i)
            continue
        with open(path, 'rb') as f:
            sum_4 = pickle.load(f)
            for key, val in sum_4.items():
                if len(list(val)) != 5:
                    unprocessed_list.append(i)
                else:
                    processed_list.append(i)

                break

    return unprocessed_list,  processed_list

def start_ev(board_ind):
    # start states
    start22 = set()
    start24 = set()
    start44 = set()
    start2 = rnd_spawn_int(0, 1)
    start4 = rnd_spawn_int(0, 2)
    for x in start2:
        start22.update(rnd_spawn_int(x,1))
        start24.update(rnd_spawn_int(x,2))
    for x in start4:
        start44.update(rnd_spawn_int(x,2))

    with open(f"polynominoes/{board_ind}/sum_4.pkl", 'rb') as f:
        sum_4 = pickle.load(f)
    with open(f"polynominoes/{board_ind}/sum_6.pkl", 'rb') as f:
        sum_6 = pickle.load(f)
    with open(f"polynominoes/{board_ind}/sum_8.pkl", 'rb') as f:
        sum_8 = pickle.load(f)


    contribution22 = 1 / 28 * 0.81
    contribution24 = 1 / 56 * 0.18
    contribution44 = 1 / 28 * 0.01


    ev = Expectations(0,0,0,0,0)
    for state_int in start22:
        ev += Expectations(sum_4[state_int][0], sum_4[state_int][1], sum_4[state_int][2], sum_4[state_int][3], sum_4[state_int][4]) * contribution22
    for state_int in start24:
        ev += Expectations(sum_6[state_int][0], sum_6[state_int][1], sum_6[state_int][2], sum_6[state_int][3], sum_6[state_int][4]) * contribution24
    for state_int in start44:
        ev += Expectations(sum_8[state_int][0], sum_8[state_int][1], sum_8[state_int][2], sum_8[state_int][3], sum_8[state_int][4]) * contribution44
    return ev


columns = ['C', 'D', 'E', 'F', 'G']
titles = ['128 rate', '256 rate', '512 rate', 'E[Score]', 'E[Sum]']

if __name__ == "__main__":
    wb = openpyxl.Workbook()
    ws = wb.active
    
    unpprocessed_list, processed_list = unprocessed()
    print("Unprocessed polynominoes:", unpprocessed_list)
    ws['A1'] = 'ID'
    ws['B1'] = 'Image'
    for i in range(5):
        ws[columns[i]+'1'] = titles[i]
        ws.column_dimensions[columns[i]].width = 12

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 7

    for i in processed_list:
        # ID
        ws['A'+str(i+2)] = i
        # image
        img_file_path = os.path.join('./polynominoes/', f"{i}/board.png")
        img = Image(img_file_path)
        img.width, img.height = (50,50)
        ws.add_image(img, f"B{i+2}")
        ev_tuple = start_ev(i).to_tuple()
        
        ws.row_dimensions[i+2].height = 35
        # info
        for j in range(5):
            ws[columns[j]+str(i+2)] = ev_tuple[j]
        
        # print(f"Expected value for polynomino {i}:", ev_tuple)
    
    wb.save('summary.xlsx')
    wb.close()